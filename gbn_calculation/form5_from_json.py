import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# from converters.hourly_con_from_xls_to_json import convert_xlsx_to_json
from useful_functions.converters.hourly_con_from_xls_to_json import convert_xlsx_to_json
from get_rmse import get_best_rmse, get_best_rmse_with_max_deviation


def forma5_from_json(data_dict):
    # Загрузка Excel-шаблона

    current_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(current_dir)
    excel_template_path = os.path.join(base_dir, 'templates', 'form_5.xlsx')
    workbook = load_workbook(excel_template_path)

    # Перебор всех листов
    for sheet_name in ["без_подстройки", "с_подстройкой", "по_вчерашнему", "20_тип_дней"]:
        sheet = workbook[sheet_name]
        gbn_list = []
        way_of_tweak = ''

        # Начальная строка и столбец для вставки данных
        start_row = 15
        date_col = 'B'
        start_col = 4  # Столбец D

        rmse_row = 2
        rrmse_row = 3
        rmse_no_tweak_col = 'B'
        rmse_with_tweak_col = 'C'
        rmse_with_tweak_yest_col = 'D'

        for_gbn_col = 'C'

        if sheet_name == "без_подстройки":
            way_of_tweak = 'min_double_rmse_no_tweak'
        elif sheet_name == "с_подстройкой":
            way_of_tweak = 'min_double_rmse_with_tweak'
        elif sheet_name == "по_вчерашнему":
            way_of_tweak = 'min_double_rmse_with_tweak_yesterday'
        elif sheet_name == "20_тип_дней" and data_dict['gbn_for_20_days']:
            way_of_tweak = 'gbn_for_20_days'
        else:
            break

        # current_data = data_dict['days_for_gbn']
        # if way_of_tweak == 'gbn_for_20_days':
        #     current_data = data_dict['gbn_for_20_days']['days'].items()

        sheet[f'{rmse_no_tweak_col}{rmse_row}'] = data_dict[way_of_tweak]['rmse_data']['doubled_rmse_no_tweak']
        sheet[f'{rmse_no_tweak_col}{rrmse_row}'] = data_dict[way_of_tweak]['rmse_data']['rrmse_no_tweak']
        sheet[f'{rmse_with_tweak_col}{rmse_row}'] = data_dict[way_of_tweak]['rmse_data'][
            'doubled_rmse_with_tweak']
        sheet[f'{rmse_with_tweak_col}{rrmse_row}'] = data_dict[way_of_tweak]['rmse_data'][
            'rrmse_with_tweak']
        sheet[f'{rmse_with_tweak_yest_col}{rmse_row}'] = data_dict[way_of_tweak]['rmse_data'][
            'doubled_rmse_with_tweak_yesterday']
        sheet[f'{rmse_with_tweak_yest_col}{rrmse_row}'] = data_dict[way_of_tweak]['rmse_data'][
            'rrmse_with_tweak_yesterday']
        gbn_list = list(data_dict[way_of_tweak]['days'].keys())



        # Заполнение данных
        for row_offset, (date, hourly_data) in enumerate(data_dict['days_for_gbn'].items()):
            # Форматирование даты
            formatted_date = datetime.strptime(date, "%Y-%m-%d").strftime("%d.%m.%Y")
            sheet[f'{date_col}{start_row + row_offset}'] = formatted_date

            if date in gbn_list:
                sheet[f'{for_gbn_col}{start_row + row_offset}'] = 'да'
            else:
                sheet[f'{for_gbn_col}{start_row + row_offset}'] = 'нет'

            # Заполнение почасовых данных
            for hour_data in hourly_data:
                for hour, value in hour_data.items():
                    col = get_column_letter(start_col + int(hour) - 1)  # Смещение для правильного столбца
                    sheet[f'{col}{start_row + row_offset}'] = round(value / 1000, 4)

    # Сохранение файла с текущей датой и временем
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    # output_filename = os.path.join("C:", os.sep, "Users", "pvsol", "Downloads", f'form_5_{current_time}.xlsx')
    output_filename = os.path.join("C:", os.sep, "Users", "psolovey.GSN07", "Downloads", f'form_5_{current_time}.xlsx')
    workbook.save(output_filename)
    print(f"Данные успешно сохранены в {output_filename}")


def form5_create(xlsx_file_name, time_zone):
    days_for_gbn = convert_xlsx_to_json(xlsx_file_name)
    data_dict = get_best_rmse(days_for_gbn, time_zone)
    # data_dict = get_best_rmse_with_max_deviation(days_for_gbn, time_zone)
    forma5_from_json(data_dict)


if __name__ == "__main__":
    form5_create("51070.xlsx", 1)
